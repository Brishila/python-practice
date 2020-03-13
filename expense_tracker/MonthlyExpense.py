from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import os.path
from Expense import Expense
from Income import Income

class MonthlyExpense:
    def __init__(self, input_type):
        self.input_type = input_type

    def input_data(self, dict_data):
        fname = 'MonthlyExpenses.xlsx'
        date = dict_data.pop('date')
        if os.path.isfile(fname):
            wb = load_workbook(fname)
        else:
            wb = Workbook()
            sheet = wb['Sheet']
            wb.remove(sheet)

        if self.input_type in wb.sheetnames:
            ws = wb[self.input_type]
        else:
            ws = wb.create_sheet(self.input_type)
            ws.append(['Date', 'Type', 'Amount'])

        for type_name, amount in dict_data.items():
            ws.append([date, type_name, amount])
            
        wb.save(fname)

if __name__ == "__main__": 
    while True:
        name_type = input("Please Enter (income/expense)").capitalize()
        if name_type == 'Income' or name_type == 'Expense':
            break
        else:
            print('Please enter valid input')
            continue
    data = {}
    while True:
        try:
            inc_type = input(f"Enter the {name_type} Type: ") 
            inc_amount = int(input("Enter the Amount: ")) 
            data[inc_type] = inc_amount
            flag = input('Wish to continue (y/n)').lower()
            if flag == 'y':
                continue
            elif flag == 'n':
                break
            else:
                print('invalid input')
                continue
        except:
            print('Invalid input')
            continue
    if name_type == 'Expense':
        class_obj = Expense(data)
        dict_value = class_obj.expense_dict
    else:
        class_obj = Income(data)
        dict_value = class_obj.income_dict

    monthly_expense = MonthlyExpense(name_type)
    monthly_expense.input_data(dict_value)